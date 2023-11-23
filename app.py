import openpyxl

workbook = openpyxl.load_workbook('Maharashtra_college_details.xlsx')

a=[]
b=[]
c=[]
columns=[]
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    a.append(sheet_name)

for sheet_name in workbook.sheetnames:
    if a[0]==sheet_name:
        a[0]=workbook[sheet_name]
        for row in a[0].iter_rows(min_row=2, values_only=True):
            b.append(row)
    if a[1]==sheet_name:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            c.append(row)
# print(a)
print(b)
print(c)
# print(columns)