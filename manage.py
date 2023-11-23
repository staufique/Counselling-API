from flask import Flask,request,jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
import openpyxl


app = Flask(__name__)

username='Your_Username'
password='Your_password'
host='@localhost'
port=5432


app.config['SQLALCHEMY_DATABASE_URI'] = f'postgresql://{username}:{password}{host}:{port}/counselling'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
ma = Marshmallow(app)

class College_Data(db.Model):
    __tablename__ = "college_data"
    id = db.Column(db.Integer(),primary_key=True,autoincrement=True)
    college_code1 = db.Column(db.String(200),nullable=False)
    college_name = db.Column(db.String(200),nullable=False)
    city = db.Column(db.String(100),nullable=False)
    state = db.Column(db.String(100),nullable=False)
    college_type = db.Column(db.String(100),nullable=False)
    courses = db.relationship('College_Course', backref='college_data', lazy=True, cascade='all, delete-orphan, delete, save-update')
    cutoff = db.relationship('College_Cutoff', backref='college_data', lazy=True, cascade='all, delete-orphan, delete, save-update')

class College_Course(db.Model):
    __tablename__ = "college_course"
    id = db.Column(db.Integer(), primary_key=True, autoincrement=True)
    college_code2 = db.Column(db.String(200),nullable=False)
    course_name = db.Column(db.String(100), nullable=False) 
    category = db.Column(db.String(100), nullable=False) 
    no_of_seats = db.Column(db.String(100),nullable=False)
    college_id = db.Column(db.Integer(),db.ForeignKey('college_data.id',ondelete='CASCADE', onupdate='CASCADE'))

class College_Cutoff(db.Model):
    __tablename__ = 'college_cutoff'
    id = db.Column(db.Integer(), primary_key=True, autoincrement=True)
    college_code3 = db.Column(db.String(200),nullable=False)
    course_name = db.Column(db.String(100), nullable=False) 
    category = db.Column(db.String(100), nullable=False)
    rank_high = db.Column(db.String(100),nullable=True)
    rank_low = db.Column(db.String(100),nullable=True)
    marks_high = db.Column(db.String(100),nullable=True)
    marks_low = db.Column(db.String(100),nullable=True)
    period = db.Column(db.String(100),nullable=False)
    college_id = db.Column(db.Integer(),db.ForeignKey('college_data.id',ondelete='CASCADE', onupdate='CASCADE'))


class College_DataSchema(ma.Schema):
    class Meta:
        model = College_Data
        fields = ["id","college_code1","college_name","city","state","college_type"]

class College_CourseSchema(ma.Schema):
    class Meta:
        model = College_Course
        fields = ["id","college_code2","course_name","category","no_of_seats","college_id"]

class College_CutoffSchema(ma.Schema):
    class Meta:
        model = College_Cutoff
        fields = ["id","college_code3","course_name","category","rank_high","rank_low","marks_high","marks_low","period","college_id"]



@app.route('/colleges',methods=['POST'])
def colleges():
    if request.method=="POST":
        file=request.files['college']
        workbook = openpyxl.load_workbook(file)
        a=[]
        sheet1=[]
        sheet2=[]
        sheet3=[]
        for sheet_name in workbook.sheetnames:
            # sheet = workbook[sheet_name]
            a.append(sheet_name)

        for sheet_name in workbook.sheetnames:
            if a[0]==sheet_name:
                a[0]=workbook[sheet_name]
                for row in a[0].iter_rows(min_row=2, values_only=True):
                    sheet1.append(row)
            if a[1]==sheet_name:
                a[1]=workbook[sheet_name]
                for row in a[1].iter_rows(min_row=2, values_only=True):
                    sheet2.append(row)
            if a[2]==sheet_name:
                a[2]=workbook[sheet_name]
                for row in a[2].iter_rows(min_row=2, values_only=True):
                    sheet3.append(row)
        for college_code1,college_name,city,state,college_type in sheet1:
            data = College_Data(college_code1=college_code1,college_name=college_name,city=city,state=state,college_type=college_type)
            db.session.add(data)
            db.session.flush()
            clg_id=data.id
            for college_code2,course_name,category,no_of_seats in sheet2:
                if college_code1==college_code2:
                    data1 = College_Course(college_code2=college_code2,course_name=course_name,category=category,no_of_seats=no_of_seats,college_id=clg_id)
                    db.session.add(data1)

            for college_code3,course_name2,category2,rank_high,rank_low,marks_high,marks_low,period in sheet3:
                if college_code1==college_code3:
                    data2 = College_Cutoff(college_code3=college_code3,course_name=course_name2,category=category2,rank_high=rank_high,rank_low=rank_low,marks_high=marks_high,marks_low=marks_low,period=period,college_id=clg_id)
                    db.session.add(data2)
        db.session.commit()
        
        
    return jsonify({'msg':'data inserted'})








@app.route('/college-details',methods=['GET','POST'])
def college_details():
    if request.method == 'GET':
        anoun = College_Data.query.all()
        schema_detail = College_DataSchema(many=True)
        return schema_detail.jsonify(anoun)
    
    if request.method == 'POST':
        college_code1 = request.json['college_code1']
        college_name = request.json['college_name']
        city = request.json['city']
        state = request.json['state']
        college_type = request.json['college_type']

        data = College_Data(college_code1=college_code1,college_name=college_name,city=city,state=state,college_type=college_type)
        db.session.add(data)
        db.session.commit()
        return jsonify({"msg":"Data Inserted"})
    return jsonify({"msg":'Select Proper Method'})


@app.route('/college-details/<int:id>',methods=['GET','PUT','DELETE'])
def college_details_crud(id):
    if request.method == 'GET':
        college_detail = College_Data.query.get_or_404(id)
        schema_detail = College_DataSchema()
        return schema_detail.jsonify(college_detail)
        # return jsonify(anoun)
    
    if request.method == 'PUT':
        college_detail = College_Data.query.get_or_404(id)
        college_code1 = request.json['college_code1']
        college_name = request.json['college_name']
        city = request.json['city']
        state = request.json['state']
        college_type = request.json['college_type']

    
        college_detail.college_code1=college_code1
        college_detail.college_name=college_name
        college_detail.city=city
        college_detail.state=state
        college_detail.college_type=college_type

        db.session.commit()
        return jsonify({"msg":"Data Updated"})
    
    if request.method == 'DELETE':
        college_detail = College_Data.query.get_or_404(id)
        db.session.delete(college_detail)
        db.session.commit()
        return jsonify({"msg":'Data Deleted Successfully'})
    return jsonify({"msg":'Select Proper Method'})









@app.route('/college-course',methods=['GET','POST'])
def college_course():
    if request.method=='GET':
        anoun = College_Course.query.all()
        schema_detail = College_CourseSchema(many=True)
        return schema_detail.jsonify(anoun)
    
    if request.method == 'POST':
        college_code2= request.json['college_code2']
        course_name=request.json['college_code2']
        category=request.json['college_code2']
        no_of_seats=request.json['college_code2']

        data = College_Course(college_code2=college_code2,course_name=course_name,category=category,no_of_seats=no_of_seats)
        db.session.add(data)
        db.session.commit()
        return jsonify({"msg":"Data Inserted"})
    return jsonify({"msg":'Select Proper Method'})


@app.route('/college-course/<int:id>',methods=['GET','PUT','DELETE'])
def college_course_crud(id):
    if request.method == 'GET':
        college_detail = College_Course.query.get_or_404(id)
        schema_detail = College_CourseSchema()
        return schema_detail.jsonify(college_detail)
    
    if request.method == 'PUT':
        college_detail = College_Course.query.get_or_404(id)
        college_code2= request.json['college_code2']
        course_name=request.json['course_name']
        category=request.json['category']
        no_of_seats=request.json['no_of_seats']

        college_detail.college_code2=college_code2
        college_detail.course_name=course_name
        college_detail.category=category
        college_detail.no_of_seats=no_of_seats
        db.session.commit()
        return jsonify({"msg":"Data Updated"})
    
    if request.method == 'DELETE':
        college_detail = College_Course.query.get_or_404(id)
        db.session.delete(college_detail)
        db.session.commit()
        return jsonify({"msg":'Data Deleted Successfully'})
    return jsonify({"msg":'Select Proper Method'})








@app.route('/college-cutoff',methods=['GET','POST'])
def college_cutoff():
    if request.method=='GET':
        anoun = College_Cutoff.query.all()
        schema_detail = College_CutoffSchema(many=True)
        return schema_detail.jsonify(anoun)
    
    if request.method == 'POST':
        college_code3=request.json['college_code3']
        course_name=request.json['course_name']
        category=request.json['category']
        rank_high=request.json['rank_high']
        rank_low=request.json['rank_low']
        marks_high=request.json['marks_high']
        marks_low=request.json['marks_low']
        period=request.json['period']

        data = College_Cutoff(college_code3=college_code3,course_name=course_name,category=category,rank_high=rank_high,rank_low=rank_low,marks_high=marks_high,marks_low=marks_low,period=period)
        db.session.add(data)
        db.session.commit()
        return jsonify({"msg":"Data Inserted"})
    return jsonify({"msg":'Select Proper Method'})


@app.route('/college-cutoff/<int:id>',methods=['GET','PUT','DELETE'])
def college_cutoff_crud(id):
    if request.method == 'GET':
        college_detail = College_Cutoff.query.get_or_404(id)
        schema_detail = College_CutoffSchema()
        return schema_detail.jsonify(college_detail)

    
    if request.method == 'PUT':
        college_detail = College_Cutoff.query.get_or_404(id)
        college_code3=request.json['college_code3']
        course_name=request.json['course_name']
        category=request.json['category']
        rank_high=request.json['rank_high']
        rank_low=request.json['rank_low']
        marks_high=request.json['marks_high']
        marks_low=request.json['marks_low']
        period=request.json['period']
      
        college_detail.college_code3=college_code3
        college_detail.course_name=course_name
        college_detail.category=category
        college_detail.rank_high=rank_high
        college_detail.rank_low=rank_low
        college_detail.marks_high=marks_high
        college_detail.marks_low=marks_low
        college_detail.period=period

        db.session.commit()
        return jsonify({"msg":"Data Updated"})
    
    if request.method == 'DELETE':
        college_detail = College_Cutoff.query.get_or_404(id)
        db.session.delete(college_detail)
        db.session.commit()
        return jsonify({"msg":'Data Deleted Successfully'})
    return jsonify({"msg":'Select Proper Method'})




if __name__ == "__main__":
    app.run(debug=True)
